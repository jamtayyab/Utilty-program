
# Mulitple Folder maker

import os
from openpyxl import load_workbook


wb=load_workbook("Provide Logos (1000by2).xlsx")
ws = wb.active
first_column = ws['B']

i = 0
for x in range(1,len(first_column)):
    name = first_column[x].value
    chr = "/"
    if chr in name:
        name = name.replace(chr, "_")
    string1 = (str(i)+". "+str(name))

    os.mkdir(string1)
    i+=1
    

